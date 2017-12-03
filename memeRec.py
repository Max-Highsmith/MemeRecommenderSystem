# -*- coding: utf-8 -*-
"""
Created on Sun Dec  3 01:45:10 2017

@author: Jon-Michael
"""

import numpy as np
from openpyxl import load_workbook
import random as rnd

from PIL import Image
import requests
from io import BytesIO
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import math



def pickFirstMeme(options):
    reduce = options[0]
    for i in range(0, len(options)):
        if options[i][1] != None:
            if options[i][1] > reduce[1]:
                reduce = options[i]
    return reduce

def nextRecommend(users, index, allMemes):
    recommend = [0, 0]
    for i in range(1, len(users[0])):
        hits = 0
        unpicked = True
        for j in range(0, len(index)):
            if i == index[j]:
                unpicked = False
        for j in range(0, len(users)):
            if users[j][i] == True:
                hits += 1
        if hits > recommend[1] and unpicked:
            recommend = [i, hits]
    giveback = allMemes[recommend[0]]
    return giveback

def getImage(url):
    response = requests.get(url)
    img = Image.open(BytesIO(response.content))
    return img
    
def getMemeNum(url, basememes):
    for i in range(0, len(basememes)):
        if basememes[i][0] == url:
            return i
    return -1
    
def userLikes(users, index):
    incrowd = []
    for i in range(0, len(users)):
        if users[i][index] == True:
            incrowd.append(users[i])
    return incrowd

def userHates(users, index):
    incrowd = []
    for i in range(0, len(users)):
        if users[i][index] != True:
            incrowd.append(users[i])
    return incrowd

def findBestFriends(users, picks):
    bffs = []
    for i in range(0, len(users)):
        verity = 0
        for j in range(0, len(picks)-1):
            if users[i][picks[j][0]] == picks[j][1] or users[i][picks[j][0]] == None:
                verity = verity + 1
        if (verity / len(picks)) >= .8:
            bffs.append(users[i])
    return bffs

def generalChop(users, picks):
    genChop = []
    broad = []
    totalpicks = len(picks)
    for i in range(0, len(users)):
        for j in range(0, len(picks)):
            if users[i][picks[j][0]] != None:
                broad.append(users[i])
    criteria = math.ceil(totalpicks/2)
    repeat = 1
    for i in range(1, len(broad)):
        if broad[i-1][0] == broad[i][0]:
            repeat = repeat + 1
        else:
            repeat = 1
        if repeat >= criteria:
            genChop.append(broad[i])
    return genChop
    
def checkpreference(user, history):
    better = False
    quota = 0
    for j in range(0, len(history)):
        if user[history[j][0]] == None or user[history[j][0]] == history[j][1]:
            quota = quota + 1
    if quota / len(history) >= .5:
        better = True
    return better
    
def noduplicates(friendlist, user):
    notpresent = True
    for i in range(0, len(friendlist)):
        if user[0] == friendlist[i][0]:
            notpresent = False
    return notpresent

def unionFriends(basefriends, newfriends, history):
    bffl = []    
    for i in range(0, len(basefriends)):
        if checkpreference(basefriends[i], history):            
            bffl.append(basefriends[i])
    for i in range(0, len(newfriends)):
        if noduplicates(basefriends, newfriends[i]):
            if checkpreference(newfriends[i], history):
                bffl.append(newfriends[i])
    return bffl



def createOptions(peerGroup, index, memeBase):
    choices = []
    expIndex = []
    for i in range(0, len(index)):
        expIndex.append(index[i])
    for i in range(0, 3):
        newChoice = nextRecommend(peerGroup, expIndex, memeBase)
        memeNum = getMemeNum(newChoice[0],memeBase)
        feedback = 0
        for j in range(0, len(peerGroup)):
            if peerGroup[j][memeNum] == True:
                feedback = feedback + 1
        choices.append([newChoice[0],memeNum, feedback])
    return choices
    

def weightedFriendChoice(unionedFr, options):
    weighted = []
    for i in range(0, len(options)):
        weight = options[i][2]
        for j in range(0, len(unionedFr)):
            if unionedFr[j][options[i][1]] == True:
                weight = weight + 1
        weighted.append([options[i][0], options[i][1], weight])
    finalchoice = [0, 0, 0]
    for i in range(0, len(weighted)):
        if weighted[i][2] > finalchoice[2]:
            for j in range(0, len(finalchoice)):
                finalchoice[j] = weighted[i][j]
    return finalchoice
    
workbook = load_workbook("MemeSurveyData.xlsx", data_only=True)
worksheet = workbook.worksheets[0]

row_count = worksheet.max_row

row = 0

data = []

for row in worksheet.iter_rows():
    holder = []
    for cell in row:
        holder.append(cell.internal_value)
    data.append(holder)

memeBaseData = []

for i in range(1, len(data)):
    memeBaseData.append([data[i][0], data[i][1], data[i][2]])

buildUsers = []

for i in range(3, len(data[0])):
    buildUsers.append([data[0][i]]);

for i in range(3, len(data[0])):
    for j in range(1, len(data)):
        buildUsers[i-3].append(data[j][i])
        
meme = pickFirstMeme(memeBaseData)

url = meme[0]
img = getImage(url)

imgplot = plt.imshow(img)
plt.show()

picks = [getMemeNum(url, memeBaseData)]

history = []

like = input("Type yes if you liked, no otherwise:")

bestfriends = []

"""
for i in range(0, 4):
    indexer = getMemeNum(url, memeBaseData)
    history.append([indexer, like])
    if like == "yes" or like == "Yes":
        friends = userLikes(buildUsers, indexer)
        stuff = nextRecommend(friends, picks, memeBaseData)
        url = stuff[0]
        img = getImage(url)
        
        imgplot = plt.imshow(img)
        plt.show()
        
        like = input("Type yes if you liked, no otherwise:")
        picks.append(getMemeNum(url, memeBaseData))
    else:
        #do other stuff
        friends = userHates(buildUsers, indexer)
        stuff = nextRecommend(friends, picks, memeBaseData)
        url = stuff[0]
        img = getImage(url)
        
        imgplot = plt.imshow(img)
        plt.show()
        
        like = input("Type yes if you liked, no otherwise:")
        picks.append(getMemeNum(url, memeBaseData))
"""

for i in range(0, 4):
    indexer = getMemeNum(url, memeBaseData)
    history.append([indexer, like])
    if like == "yes" or like == "Yes":
        friends = userLikes(buildUsers, indexer)
        stuff = weightedFriendChoice(
            unionFriends(bestfriends, friends, history),
            createOptions(friends, picks, memeBaseData)        
        )
        url = stuff[0]
        img = getImage(url)
        
        imgplot = plt.imshow(img)
        plt.show()
        
        like = input("Type yes if you liked, no otherwise:")
        picks.append(getMemeNum(url, memeBaseData))
    else:
        #do other stuff
        friends = userHates(buildUsers, indexer)
        stuff = weightedFriendChoice(
            unionFriends(bestfriends, friends, history),
            createOptions(friends, picks, memeBaseData)        
        )
        url = stuff[0]
        img = getImage(url)
        
        imgplot = plt.imshow(img)
        plt.show()
        
        like = input("Type yes if you liked, no otherwise:")
        picks.append(getMemeNum(url, memeBaseData))
"""
Note to Self:

- create a ban list
- create a no option
"""

